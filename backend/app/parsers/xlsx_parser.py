"""XLSX parser — extracts sheet data via openpyxl (fallback: pandas)."""
from .base import BaseParser, ParsedContent
from ..config import get_settings


class XlsxParser(BaseParser):
    supported_extensions = [".xlsx", ".xlsm", ".xls"]

    def parse(self, filepath: str) -> ParsedContent:
        result = self._make_result(filepath, "xlsx")
        result.extraction_method = "openpyxl"
        max_chars = get_settings().max_chars_per_file
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            result.sheet_names = wb.sheetnames
            parts = []
            for sname in wb.sheetnames[:5]:
                ws = wb[sname]
                parts.append(f"=== Sheet: {sname} ===")
                count = 0
                for row in ws.iter_rows(values_only=True):
                    if count >= 100:
                        parts.append("...[more rows omitted]")
                        break
                    row_text = " | ".join(str(v) for v in row if v is not None)
                    if row_text.strip():
                        parts.append(row_text)
                        count += 1
            combined = "\n".join(parts)
            result.text_content = combined[:max_chars]
        except Exception:
            try:
                import pandas as pd
                result.extraction_method = "pandas"
                xl = pd.ExcelFile(filepath)
                result.sheet_names = xl.sheet_names
                parts = []
                for s in xl.sheet_names[:5]:
                    df = xl.parse(s, nrows=80)
                    parts.append(f"=== Sheet: {s} ===\n{df.to_string(max_rows=80)}")
                result.text_content = "\n".join(parts)[:max_chars]
            except Exception as exc2:
                result.error = str(exc2)
                result.text_content = f"[Could not read Excel: {exc2}]"
        return result
